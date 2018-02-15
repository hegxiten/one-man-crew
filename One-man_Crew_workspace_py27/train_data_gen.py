from __future__ import division
import networkx as nx
import matplotlib as plt
import numpy as np
import random
from openpyxl import Workbook


d = 360
T = 24
avg_speed = 60
speed_sigma = 5
train_number = 25
time_step = 1/60
absolute_directory = "D:\\Users\\Hegxiten\\workspace\\CEE512\\"

train_list = []

def gen_symb(dirac, train_idx):
    symb = ""
    if dirac == 1:
        symb += "A"
    elif dirac == -1:
        symb += "B"
    if len(str(train_idx)) == 1:
        symb += "00"
        symb += str(train_idx)
    elif len(str(train_idx)) ==2:
        symb += "0"
        symb += str(train_idx)
    else:
        symb += str(train_idx)
    return symb

def depart_a_train(dirac = 1):
    train = {}
    
    if dirac == -1:
        (dep_mp, arr_mp) = (d, 0)
    if dirac == 1:
        (dep_mp, arr_mp) = (0, d)
        
    train["dep_arr_tuple"] = (dep_mp, arr_mp)
    
    dep_t = np.random.uniform(0, T)
    train["dep_time"] = dep_t
    
    avg_spd = np.random.normal(avg_speed, speed_sigma)
    if dirac == -1:
        train["avg_speed"] = -avg_spd
    if dirac == +1:
        train["avg_speed"] = avg_spd
    
    idx = len(train_list)
    train["symb"] = gen_symb(dirac, idx)
    
    t_mp_list = []
    # t_mp tuple is the pair of (time mile-post) for a train to determine the train status
    t_mp_list.append((dep_t,dep_mp))
    train["time_location"] = t_mp_list
    train_list.append(train)

def dirac_gen():
    dirac = random.randrange(-1,3,2)
    return int(dirac)
    
for i in range(train_number):
    dirac = dirac_gen()
    depart_a_train(dirac)

for i in range(len(train_list)):
    
    time_delta = time_step
    current_t = train_list[i]['time_location'][-1][0]
    current_mp = train_list[i]['time_location'][-1][1]
    spd = train_list[i]['avg_speed']
    dep_mp = train_list[i]['dep_arr_tuple'][0]
    
    while current_t + time_delta <= T and abs(current_mp + spd * time_delta - dep_mp) <= d:

        current_mp += spd * time_delta
        current_t += time_delta
        train_list[i]['time_location'].append((current_t,current_mp))
    
    last_arr_t = train_list[i]['time_location'][-1][0]
    
    if T - last_arr_t <= time_delta: 
        arr_mp = train_list[i]['time_location'][-1][1] + spd * (T - last_arr_t)
        arr_t = T
        train_list[i]['time_location'].append((arr_t,arr_mp))

    if spd < 0:
        arr_t = train_list[i]['time_location'][-1][0] + (train_list[i]['time_location'][-1][1]-0) / -spd
        if arr_t > T:
            continue
        arr_mp = 0
        train_list[i]['time_location'].append((arr_t,arr_mp))

    elif spd >= 0:
        arr_t = train_list[i]['time_location'][-1][0] + (d-train_list[i]['time_location'][-1][1]) / spd 
        if arr_t > T:
            continue
        arr_mp = d
        train_list[i]['time_location'].append((arr_t,arr_mp))

if __name__ == '__main__':
    
    wb = Workbook()
    ws1 = wb.create_sheet("trains_data")

    ws1['A1'] = 'Number of Trains'
    ws1['B1'] = len(train_list)
    
    for i in range(len(train_list)):
        ws1.cell(column = 2*i+1, row = 2, value = "Train Symbol")
        ws1.cell(column = 2*i+2, row = 2, value = "Train Index")
        ws1.cell(column = 2*i+1, row = 3, value = train_list[i]["symb"])
        ws1.cell(column = 2*i+2, row = 3, value = i)
        ws1.cell(column = 2*i+1, row = 4, value = "Time")
        ws1.cell(column = 2*i+2, row = 4, value = "Mile Post")
        for r in range(len(train_list[i]["time_location"])):
            ws1.cell(column = 2*i+1, row = 5+r, value = train_list[i]["time_location"][r][0])
            ws1.cell(column = 2*i+2, row = 5+r, value = train_list[i]["time_location"][r][1])
    wb.save(absolute_directory + "data.xlsx")
    
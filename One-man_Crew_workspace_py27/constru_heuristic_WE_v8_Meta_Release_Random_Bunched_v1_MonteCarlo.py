from __future__ import division
import time
from types import NoneType
startt = time.time()
import matplotlib.pyplot as plt
import openpyxl
import networkx as nx
import numpy as np
import random
import itertools as it
from openpyxl import Workbook, load_workbook

absolute_directory = "D:\\Users\\Hegxiten\\workspace\\CEE512\\"
fn = "data.xlsx"
#cond_start = int(raw_input("Where does conductor start: "))
#shift_time = float(raw_input("Shift time limitation in min: "))



def data_init(sample = True, WE_number = 20):

    if sample:
        x = []
        t = []
        WE_ind_list = []
        WExstr =    '245.9331686775067  344.1195289509466  216.1801348173958  153.2589438000065  239.4776119326927  154.105498325298  152.2671545603453  101.7127010241457  87.92512278700025  21.75185327669039  38.06574323420818  50.51170408230033  287.8526463459218  179.4527895326966  328.0904535900797  347.1233252071834  186.2631559273912  189.40653929477  193.8714176040145  265.5700154133025' #distance
        WEtstr =    '3.090111865271252  6.118171955609674  6.590435411451304  7.757102078117966  9.181491241968619  9.61029015573338  9.643623489066716  10.56029015573343  10.81029015573344  12.30227721014547  12.60227721014549  13.28691823722243  14.00743248814362  15.20227721014562  17.93561054347888  18.28561054347886  22.01555417655099  22.24675508804026  22.82631686062795  23.51342175470685' #time
        WEindex =   '1  2  3  4  5  6  7  8  9  10  11  12  13  14  15  16  17  18  19  20'
        
        WExstr = '126.6644347    133.9647889    202.5735652    81.35277241    70.90657674    123.8117143    249.1785663    142.5841096    233.8498302    34.83084512    258.6229286    283.8504342    4.664852002    94.09322035    190.2300003    76.32967966    176.8538531    104.1090435    153.9908383    94.50654568    176.5765181    259.975164    51.19597459    256.8350699    138.2533771    150.932071    261.6246317    215.8613817    204.5801653    193.8023947    139.9843049    118.9978267    133.8622901    243.911081    294.6738265    139.8157851    4.996791243    42.35198115    267.3515073    167.9894702    144.6819721    239.8440172    84.63630197    205.8707526    124.7179123    175.539346    99.93582487    42.92852018    292.2977799    10.96481426    54.96470368    212.863307    276.4125474    166.6796155    148.572871    67.95636091    73.90262928    94.10551232    60.65986548    169.1634004    204.3729906    241.8446962    232.2056342    110.9287656    120.2606201    6.163565323    3.368774613    0.9332287    269.2720068    230.6800187    26.53005827    84.15836475    210.0739215    170.2358316    209.8899334    145.2067925    213.7184144    110.5336582    193.2436585    176.4665881    72.9039451    295.6243872    300    123.0488612    87.94352588    209.7236777    153.9011703    35.17074703    233.5683513    233.3398078    2.592988439    244.2941078    144.9928197    4.033522014    39.399232    278.8517191    1.538969606    244.2941078    166.3481278    63.60806638' #distance
        WEtstr = '8.798570808    7.072111538    2.252202632    2.228952632    6.792078218    5.213177145    4.388244719    8.215310798    7.426096437    5.222111538    10.78180025    6.045619299    11.08173105    5.809945316    8.122111538    11.56417933    4.028952632    3.956558044    8.415310798    4.123224711    5.554911385    5.595619299    6.476611983    7.809429771    4.963177145    4.926611983    4.188244719    7.126096437    5.104911385    8.188778205    10.46960996    5.296510479    3.368869298    11.04846692    10.88197746    7.942078218    3.609429771    1.684696068    1.123224711    8.198570808    7.272111538    4.538244719    4.168869298    7.648570808    6.388244719    5.571578052    5.192763104    8.736276627    2.296510479    5.573224711    4.442763104    7.076096437    1.052202632    2.835535965    6.004911385    4.659429771    6.842078218    7.365310798    9.052943294    4.643278649    11.76513359    7.559429771    8.905444871    5.376096437    2.962285966    5.656558044    8.338244719    7.986276627    1.089891377    3.363177145    6.859945316    11.69751266    9.398644131    4.626611983    11.66513359    3.384696068    9.175411551    2.778952632    7.831904142    3.901362734    6.825411551    10.89864413    2.163177145    5.359945316    4.992763104    9.108744885    6.092763104    6.331977464    3.313177145    1.752202632    5.502202632    9.998644131    4.846510479    1.051362734    6.659945316    10.41513359    7.329843812    9.998644131    8.631977464    11.34751266' #time
        WEindex = '1    2    3    4    5    6    7    8    9    10    11    12    13    14    15    16    17    18    19    20    21    22    23    24    25    26    27    28    29    30    31    32    33    34    35    36    37    38    39    40    41    42    43    44    45    46    47    48    49    50    51    52    53    54    55    56    57    58    59    60    61    62    63    64    65    66    67    68    69    70    71    72    73    74    75    76    77    78    79    80    81    82    83    84    85    86    87    88    89    90    91    92    93    94    95    96    97    98    99    100'        
        
        WExstr = '126.6644347    133.9647889    202.5735652    81.35277241    70.90657674    123.8117143    249.1785663    142.5841096    233.8498302    34.83084512    258.6229286    283.8504342    4.664852002    94.09322035    190.2300003    76.32967966    176.8538531    104.1090435    153.9908383    94.50654568    176.5765181    259.975164    51.19597459    256.8350699    138.2533771    150.932071    261.6246317    215.8613817    204.5801653    193.8023947    139.9843049    118.9978267    133.8622901    243.911081    294.6738265    139.8157851    4.996791243    42.35198115    267.3515073    167.9894702    144.6819721    239.8440172    84.63630197    205.8707526    124.7179123    175.539346    99.93582487    42.92852018    292.2977799    10.96481426    54.96470368    212.863307    276.4125474    166.6796155    148.572871    67.95636091    73.90262928    94.10551232    60.65986548    169.1634004    204.3729906    241.8446962    232.2056342    110.9287656    120.2606201    6.163565323    3.368774613    0.9332287    269.2720068    230.6800187    26.53005827    84.15836475    210.0739215    170.2358316    209.8899334    145.2067925    213.7184144    110.5336582    193.2436585    176.4665881    72.9039451    295.6243872    300    123.0488612    87.94352588    209.7236777    153.9011703    35.17074703    233.5683513    233.3398078    2.592988439    244.2941078    144.9928197    4.033522014    39.399232    278.8517191    1.538969606    244.2941078    166.3481278    63.60806638' #distance
        WEtstr = '8.798570808    7.072111538    2.252202632    2.228952632    6.792078218    5.213177145    4.388244719    8.215310798    7.426096437    5.222111538    10.78180025    6.045619299    11.08173105    5.809945316    8.122111538    11.56417933    4.028952632    3.956558044    8.415310798    4.123224711    5.554911385    5.595619299    6.476611983    7.809429771    4.963177145    4.926611983    4.188244719    7.126096437    5.104911385    8.188778205    10.46960996    5.296510479    3.368869298    11.04846692    10.88197746    7.942078218    3.609429771    1.684696068    1.123224711    8.198570808    7.272111538    4.538244719    4.168869298    7.648570808    6.388244719    5.571578052    5.192763104    8.736276627    2.296510479    5.573224711    4.442763104    7.076096437    1.052202632    2.835535965    6.004911385    4.659429771    6.842078218    7.365310798    9.052943294    4.643278649    11.76513359    7.559429771    8.905444871    5.376096437    2.962285966    5.656558044    8.338244719    7.986276627    1.089891377    3.363177145    6.859945316    11.69751266    9.398644131    4.626611983    11.66513359    3.384696068    9.175411551    2.778952632    7.831904142    3.901362734    6.825411551    10.89864413    2.163177145    5.359945316    4.992763104    9.108744885    6.092763104    6.331977464    3.313177145    1.752202632    5.502202632    9.998644131    4.846510479    1.051362734    6.659945316    10.41513359    7.329843812    9.998644131    8.631977464    11.34751266' #time
        WEindex = '1    2    3    4    5    6    7    8    9    10    11    12    13    14    15    16    17    18    19    20    21    22    23    24    25    26    27    28    29    30    31    32    33    34    35    36    37    38    39    40    41    42    43    44    45    46    47    48    49    50    51    52    53    54    55    56    57    58    59    60    61    62    63    64    65    66    67    68    69    70    71    72    73    74    75    76    77    78    79    80    81    82    83    84    85    86    87    88    89    90    91    92    93    94    95    96    97    98    99    100'          
        
        WExstr = '0.9332287    1.538969606    2.592988439    3.368774613    4.033522014    4.664852002    4.996791243    6.163565323    10.96481426    26.53005827    34.83084512    35.17074703    39.399232    42.35198115    42.92852018    51.19597459    54.96470368    60.65986548    63.60806638    67.95636091' #distance
        WEtstr = '7.986276627    7.329843812    5.502202632    8.338244719    1.051362734    11.08173105    3.609429771    5.656558044    5.573224711    6.859945316    5.222111538    6.331977464    6.659945316    1.684696068    8.736276627    6.476611983    4.442763104    9.052943294    11.34751266    4.659429771' #time
        WEindex = '1    2    3    4    5    6    7    8    9    10    11    12    13    14    15    16    17    18    19    20'     
        
        x = [float(n) for n in WExstr.split()]
        t = [float(n) for n in WEtstr.split()]
        WE_ind_list = [int(n) for n in WEindex.split()]
    if not sample:
        x1 = []
        t1 = []
        WE_ind_list1 = []
        wb = load_workbook(absolute_directory + fn)
        ws = wb['trains_data']
        n_cell = ws['B1']
        n_trains = int(n_cell.value)
        col_ind = -1
        WE_i = 0
        for col in ws.iter_cols(min_row = 5, max_col = 2*n_trains):
            col_ind += 1
            for cell in col:
                if type(cell.value) is not NoneType:
                    if col_ind % 2 == 0:
                        WE_i += 1
                        t1.append(cell.value)
                        WE_ind_list1.append(WE_i)
                    if col_ind % 2 == 1:
                        x1.append(cell.value)
                else:
                    break

        x = []
        t = []
        WE_ind_list = []
        for n in range(WE_number):
            idx = random.randint(0,len(x1)-1)
            x.append(x1.pop(idx))
            t.append(t1.pop(idx))
        
        tx = zip(t,x)
        tx.sort()
        
        t = [a for (a,b) in tx]
        x = [b for (a,b) in tx]
        for n in range(len(t)):
            WE_ind_list.append(n+1)

    w_e_dict = {}

    node_list = []
    
    for i in WE_ind_list: # connects distance and time as a coordinate
        node_list.append((i,
                          {'property'   :'WE',
                           'label'      :str(i),
                           'coord'      :(x[WE_ind_list.index(i)],t[WE_ind_list.index(i)])
                           }
                          ))
    
    
    
    return node_list

def graph_create(input, terminal_loc = [0], shift_hr = 24):
    
    G = nx.DiGraph() #creates graph
    
    G.add_nodes_from(input) #adds nodes from xy
    terminal_begin = set()
    terminal_end = set()
    for ind, t in enumerate(terminal_loc):
        G.add_node('T'+str(ind)+'Begin', 
                   {'coord'     :(t,0),
                    'property'  :'T',
                    'label'     :'T'+str(ind)+'Begin'}
                   )
        G.add_node('T'+str(ind)+'End', 
                   {'coord'     :(t,shift_hr),
                    'property'  :'T',
                    'label'     :'T'+str(ind)+'End'}
                   )
        
        terminal_begin.add('T'+str(ind)+'Begin')
        terminal_end.add('T'+str(ind)+'End')
    for (n,m) in it.permutations(G.nodes(),2):
        if n in terminal_end:
            continue
        else:
            if G.node[n]['coord'][1] < G.node[m]['coord'][1]:
                if not ((n in terminal_begin) == (m in terminal_end) == True):
    
                    G.add_edge(n, m,
                               dist     = abs(G.node[n]['coord'][0] - G.node[m]['coord'][0]), 
                               delt_t   = abs(G.node[n]['coord'][1] - G.node[m]['coord'][1]),
                               spd      = round(abs(G.node[n]['coord'][0] - G.node[m]['coord'][0]) / abs(G.node[n]['coord'][1] - G.node[m]['coord'][1]),2)
                               #speed in miles per hour
                               )
                    
    
    print G.edges()
    return G


def transfer_graph_to_solution_dic(G):
    G_c = G.copy()
    terminal_begin_list = []
    terminal_end_list = []
    for n in G_c.nodes():
        if G_c.node[n]['property'] == 'T':
            if n[2] == 'B':
                terminal_begin_list.append(n)
            if n[2] == 'E':
                terminal_end_list.append(n)

    terminal_begin_list.sort()
    terminal_end_list.sort()
    solution_dic = {}

    for i in range(len(terminal_begin_list)):
        solution_dic[terminal_begin_list[i]] = []
        for path in nx.all_simple_paths(G_c, source = terminal_begin_list[i], target=terminal_end_list[i]):
            solution_dic[terminal_begin_list[i]].append(path)
    
    return solution_dic

        
        
def constr_heu_srch(n_ind, G, dist_first = True):
    term_end_list = []
    term_begin_list = []
    for n in G.nodes():
        if G.node[n]['property'] == 'T':
            if G.out_degree(n) == 0:
                term_end_list.append(n)
            if G.in_degree(n) == 0:
                term_begin_list.append(n)        
                
    candi_list = []
    candi_dist_list = []
    candi_delt_list = [] 
    next_node = [None, None]
    if G.successors(n_ind):
        for nb in G.successors(n_ind):
            if nb in term_end_list:
                continue
            else:
                candi_list.append(nb)
                candi_dist_list.append(G.edge[n_ind][nb]['dist'])
                candi_delt_list.append(G.edge[n_ind][nb]['delt_t'])
        if dist_first == True:
            if candi_dist_list:
                next_node[0] = candi_list[candi_dist_list.index(min(candi_dist_list))]
                next_node[1] = G.edge[n_ind][nb]['dist'] / G.edge[n_ind][nb]['delt_t'] 
        if dist_first == False:
            if candi_delt_list:
                next_node[0] = candi_list[candi_delt_list.index(min(candi_delt_list))]
                next_node[1] = G.edge[n_ind][nb]['dist'] / G.edge[n_ind][nb]['delt_t'] 
    return next_node



def construct_heu_feasi_sol(G, max_speed = None, dist_first = True, rand = False):   
    solution = {}
    traversed_nodes = set()
    out_only_nodes =set()
    in_only_nodes = set()
    island_nodes = set()
    terminal_nodes = set()
    term_begin_list = []
    term_end_list = []
    shift_hr = 0
    for n in G.nodes():
        if G.node[n]['property'] == 'T':
            if n[2] == 'E':
                term_end_list.append(n)
                shift_hr = G.node[n]['coord'][1]
            if n[2] == 'B':
                term_begin_list.append(n)  
    term_end_list.sort()
    term_begin_list.sort()
    for ind in range(len(term_begin_list)):
        solution[term_begin_list[ind]] = []
        
    G_c = G.copy()
    
    if max_speed:
        for (e1,e2) in G_c.edges():
            if G_c.edge[e1][e2]['spd'] > max_speed:
                G_c.remove_edge(e1,e2)
    
    for n in G_c.nodes():
        if G_c.node[n]['property'] == 'WE':
            if G_c.node[n]['coord'][1] > shift_hr:
                G_c.remove_node(n)

    
    for n in G_c.nodes():
        if G_c.in_degree(n) == 0 and G_c.in_degree(n) != G_c.out_degree(n):
            out_only_nodes.add(n)
        if G_c.out_degree(n) == 0 and G_c.in_degree(n) != G_c.out_degree(n):
            in_only_nodes.add(n)
        if G_c.in_degree(n) == G_c.out_degree(n) == 0:
            island_nodes.add(n)
    
    
    if rand:        
        a = range(len(term_begin_list))
        random.shuffle(a)
        term_begin_list = [term_begin_list[i] for i in a]
        term_end_list = [term_end_list[i] for i in a]
    
    #-----------------------------------------------#
    indicator = 0
    while set(G_c.nodes())-set(term_begin_list)-set(term_end_list)!= traversed_nodes and indicator < len(term_begin_list):
        if rand:        
            a = range(len(term_begin_list))
            random.shuffle(a)
            term_begin_list = [term_begin_list[i] for i in a]
            term_end_list = [term_end_list[i] for i in a]
        indicator = 0 
        for ind in range(len(term_begin_list)):
            if not nx.has_path(G_c, source=term_begin_list[ind], target=term_end_list[ind]):
                indicator += 1
    
            for tn in traversed_nodes:
                if tn in G_c.nodes():
                    G_c.remove_node(tn)
            path = [term_begin_list[ind]]
            flag_t = path[-1][:2]
            
            while (len(G.successors(path[-1])) > 0): #list of neighbors
                
                flag = path[-1]
                next_heu_candi = constr_heu_srch(path[-1], G = G_c, dist_first=dist_first)
                if (next_heu_candi[0], term_end_list[ind]) in G.edges():
                    if next_heu_candi[0] == term_end_list[ind]:
                        path.append(next_heu_candi[0])
                    if (path[-1], next_heu_candi[0]) in G_c.edges():
                        path.append(next_heu_candi[0])
                else:
                    if (path[-1], term_end_list[ind]) in G_c.edges():
                        path.append(term_end_list[ind])
                if not next_heu_candi[0]:
                    if ([path[-1]],[term_end_list[ind]]) in G.edges():
                        if G.edge[path[-1]][term_end_list[ind]]['spd'] <= max_speed:
                            path.append(term_end_list[ind])
                
                if type(path[-1]) is str:
                    if path[-1][:2] == flag_t:
                        break
                if path[-1] == flag:
                    break
            
            if path[-1] != term_end_list[ind]:
                flag_2 = 0
                while flag_2 == 0 and len(path)>1:
                    path.pop(-1)
                    if (path[-1], term_end_list[ind]) in  G_c.edges():
                        path.append(term_end_list[ind])
                        flag_2 = 1

            


            if path[-1] == path[0]:
                if constr_heu_srch(path[-1], G_c, dist_first=dist_first)[0]:
                    G_c.remove_edge(path[-1],constr_heu_srch(path[-1], G_c, dist_first=dist_first)[0])
            if path[0] in term_begin_list and path[-1] in term_end_list and len(path)>2 and path[0][1] == path[-1][1]:
                for t_n in path[1:-1]:
                    if t_n not in terminal_nodes:
                        traversed_nodes.add(t_n)            

                solution[term_begin_list[ind]].append(path)
                    
        
        #-----------------------------------------------#

            
    visited = set()
    t_c_number = {}
    
    for t,p in solution.iteritems():
        t_c_number[t] = len(p)
        for n in p:
            for i in n:
                visited.add(i)

    t_nmbr = 0
    for t in t_c_number:
        t_nmbr += t_c_number[t]
    
    return_list = [solution, 
                   {'terminals':len(solution.keys()),
                    'Title':'Construction Heuristic Result', 
                    'conductor number':t_nmbr, 
                    'remained W.E.':len(G.nodes())-len(visited)
                    }
                   ]
    print solution
    return return_list

def path_release(G, sol_dic, max_speed = None, release_condition = 1):
    G_c = G.copy()
    
    path_list = []
    edge_list = []
    for k,v in sol_dic.iteritems():
        for p in v:
            path_list.append(p)
            for e in [(p[n], p[n+1]) for n in range(len(p)-1)]:
                edge_list.append(e)
            
    for e in G_c.edges():
        if e not in edge_list:
            G_c.remove_edge(e[0],e[1])
    #--------------------------------------#                
    n_of_cond = [-1,0]
    traversed_n = []
    while n_of_cond[-1] != n_of_cond[-2]:
        current_sol = transfer_graph_to_solution_dic(G_c)
        #------------------------------------------------#
        p_l = []
        for k,v in current_sol.iteritems():
            p_l.extend(v)
        n_of_cond.append(len(p_l))
    
        #------------------------------------------------
        path_list = []
        edge_list = []
        for k,v in current_sol.iteritems():
            for p in v:
                path_list.append(p)
                for e in [(p[n], p[n+1]) for n in range(len(p)-1)]:
                    edge_list.append(e)
                
        for e in G_c.edges():
            if e not in edge_list:
                G_c.remove_edge(e[0],e[1])
        
        all_edge = {}
        release_candi_path_dic = {}
        release_candi_we_dic = {}
        insert_candi_edge_list = []
        
        for p in path_list:
            release_candi_path_dic[p[0]] = []
            if len(p[1:-1]) <= release_condition:
                release_candi_path_dic[p[0]].append(p)
                for n in p[1:-1]:
                    release_candi_we_dic[n] = p
        #------------------------------------------------
        edge2release = []
        for n, c_p in release_candi_we_dic.iteritems():
            for e2r in [(c_p[indx], c_p[indx+1]) for indx in range(len(c_p)-1)]:
                edge2release.append(e2r)
        
        
        for n, c_p in release_candi_we_dic.iteritems():

            if n not in traversed_n:
                traversed_n.append(n)
            else:
                continue
            cut_p = c_p[1:-1]

            edge2insert = []

            for e2i in G_c.edges():
                if e2i not in edge2release:
                    edge2insert.append(e2i)
            ct = 0
            insertion_dic = {}
            insertion_dic[n] = {}
            for e in edge2insert:
                
                if G.has_edge(e[0],n) and G.has_edge(n,e[1]):
                    if G.edge[e[0]][n]['spd'] <= max_speed and G.edge[n][e[1]]['spd'] <= max_speed:
                        dist_cost = G.edge[e[0]][n]['dist'] + G.edge[n][e[1]]['dist'] - G.edge[e[0]][e[1]]['dist'] - G.edge[G_c.predecessors(n)[0]][n]['dist'] - G.edge[n][G_c.successors(n)[0]]['dist']   
                        insertion_dic[n][e]=dist_cost
            
            if insertion_dic[n]:
                ct+=1
                
            if ct == len(cut_p):
                for k in insertion_dic:
                    
                    dist_l = []
                    d_e_l = []
                    for ed,v in insertion_dic[k].iteritems():
                        dist_l.append(v)
                        d_e_l.append(ed)
                    sorted_d_e_l = [x for _, x in sorted(zip(dist_l, d_e_l))]
                    d_e = sorted_d_e_l[0]
                
                    print 'insert',k, 'to', d_e
                    if G_c.predecessors(k)[0] != e[0] and G_c.successors(k)[0] != e[1]:
                        print 'removed', d_e[0], d_e[1]
                        G_c.remove_edge(d_e[0], d_e[1])
                        print 'removed', G_c.predecessors(k)[0],k
                        G_c.remove_edge(G_c.predecessors(k)[0],k)
                        print 'removed', k, G_c.successors(k)[0]
                        G_c.remove_edge(k, G_c.successors(k)[0])
                        print 'linked', d_e[0], k    
                        G_c.add_edge(d_e[0], k, attr_dict = G.get_edge_data(d_e[0], k))
                        print 'linked', k, d_e[1]
                        G_c.add_edge(k, d_e[1], attr_dict = G.get_edge_data(k, d_e[1]))

                    elif  G_c.successors(k)[0]==e[1]:
                        print 'removed', d_e[0], d_e[1]
                        G_c.remove_edge(d_e[0], d_e[1])
                        print 'removed', G_c.predecessors(k)[0],k
                        G_c.remove_edge(G_c.predecessors(k)[0],k)
                        print 'linked', d_e[0], k
                        G_c.add_edge(d_e[0], k, attr_dict = G.get_edge_data(d_e[0], k))
                        print 'linked', k, e[1]
                        G_c.add_edge(k, d_e[1], attr_dict = G.get_edge_data(k, d_e[1]))
                        
                    elif G_c.predecessors(k)[0]==e[0]:
                        print 'removed', d_e[0], d_e[1]
                        G_c.remove_edge(d_e[0], d_e[1])
                        print 'removed', k, G_c.successors(k)[0]
                        G_c.remove_edge(k, G_c.successors(k)[0])
                        print 'linked', k, d_e[1]
                        G_c.add_edge(k, d_e[1], attr_dict = G.get_edge_data(k, d_e[1]))
                        print 'linked', d_e[0], k 
                        G_c.add_edge(d_e[0], k, attr_dict = G.get_edge_data(d_e[0], k))
    print n_of_cond
    solution = transfer_graph_to_solution_dic(G_c)
    visited = set()
    t_c_number = {}
    
    for t,p in solution.iteritems():
        t_c_number[t] = len(p)
        for n in p:
            for i in n:
                visited.add(i)

    t_nmbr = 0
    for t in t_c_number:
        t_nmbr += t_c_number[t]
    
    return_list = [solution, 
                   {'terminals':len(solution.keys()),
                    'Title':'Single Roaming Conductor Deletion', 
                    'conductor number':t_nmbr, 
                    'remained W.E.':len(G.nodes())-len(visited)
                    }
                   ]
    return return_list
    



def meta_heu(G, sol_dic, max_speed = None):
    G_c = G.copy()
    path_list = []
    edge_list = []
    for k,v in sol_dic.iteritems():
        for p in v:
            path_list.append(p)
            for e in [(p[n], p[n+1]) for n in range(len(p)-1)]:
                edge_list.append(e)
            
    for e in G_c.edges():
        if e not in edge_list:
            G_c.remove_edge(e[0],e[1])
    avg_speed = [-1, 0]
    ct = 0
    while avg_speed[-1] != avg_speed[-2]:
        ct += 1
        print 'number of meta-heuristic traverse time: ' + str(ct)

        new_total_dist = 0
        new_total_time = 0
        for e in G_c.edges():
            new_total_dist += G_c.edge[e[0]][e[1]]['dist']
            new_total_time += G_c.edge[e[0]][e[1]]['delt_t']
            
        new_avg_speed = new_total_dist/new_total_time
        avg_speed.append(new_avg_speed)
        for (n,m) in it.combinations(G_c.nodes(),2):
        
            if G_c.in_degree(n) == 1 and G_c.out_degree(n) == 1 and G_c.in_degree(m) == 1 and G_c.out_degree(m) == 1:
                (pre_n, suc_n) = (G_c.predecessors(n)[0], G_c.successors(n)[0])
                (pre_m, suc_m) = (G_c.predecessors(m)[0], G_c.successors(m)[0])
                dist_n_ori = G_c.edge[pre_n][n]['dist'] + G_c.edge[n][suc_n]['dist']
                dist_m_ori = G_c.edge[pre_m][m]['dist'] + G_c.edge[m][suc_m]['dist']
                [spd_pre_m_n,spd_n_suc_m,spd_pre_n_m,spd_m_suc_n] = [0,0,0,0]
                if (pre_m, n) in G.edges():
                    spd_pre_m_n = G.edge[pre_m][n]['spd']
                if (n, suc_m) in G.edges():
                    spd_n_suc_m = G.edge[n][suc_m]['spd']
                if (pre_n, m) in G.edges():
                    spd_pre_n_m = G.edge[pre_n][m]['spd']
                if (m, suc_n) in G.edges():
                    spd_m_suc_n = G.edge[m][suc_n]['spd']
                
                if spd_pre_m_n and spd_n_suc_m and spd_pre_n_m and spd_m_suc_n:
                    if max_speed:
                        if spd_pre_m_n <= max_speed and spd_n_suc_m <= max_speed and spd_pre_n_m <= max_speed and spd_m_suc_n <= max_speed:
                            if (pre_m, n) in G.edges() and (n, suc_m) in G.edges() and (pre_n, m) in G.edges() and(m, suc_n) in G.edges():
                                dist_n_prop = G.edge[pre_m][n]['dist'] + G.edge[n][suc_m]['dist']
                                dist_m_prop = G.edge[pre_n][m]['dist'] + G.edge[m][suc_n]['dist']
                                if dist_n_prop + dist_m_prop < dist_n_ori +dist_m_ori:
                                    print str()+'meta-heuristic swapping executed between: ',(n,m)
                                    G_c.add_edge(pre_m, n, attr_dict = G.get_edge_data(pre_m, n))
                                    G_c.add_edge(n, suc_m, attr_dict = G.get_edge_data(n, suc_m))
                                    G_c.add_edge(pre_n, m, attr_dict = G.get_edge_data(pre_n, m))
                                    G_c.add_edge(m, suc_n, attr_dict = G.get_edge_data(m, suc_n))
                                    
                                    G_c.add_edge(pre_m, n, attr_dict = G.get_edge_data(pre_m, n))
                                    G_c.add_edge(n, suc_m, attr_dict = G.get_edge_data(n, suc_m))
                                    G_c.add_edge(pre_n, m, attr_dict = G.get_edge_data(pre_n, m))
                                    G_c.add_edge(m, suc_n, attr_dict = G.get_edge_data(m, suc_n))
                                    
    
                                    if suc_n == suc_m:
                                        G_c.remove_edge(pre_m,m)
                                        G_c.remove_edge(pre_n,n)
                                    else:
                                        if pre_n == pre_m:
                                            G_c.remove_edge(m,suc_m)
                                            G_c.remove_edge(n,suc_n)
                                        else:
                                            G_c.remove_edge(pre_m,m)
                                            G_c.remove_edge(pre_n,n)
                                            G_c.remove_edge(n,suc_n)
                                            G_c.remove_edge(m,suc_m)
                    if not max_speed:
                        if (pre_m, n) in G.edges() and (n, suc_m) in G.edges() and (pre_n, m) in G.edges() and(m, suc_n) in G.edges():
                            dist_n_prop = G.edge[pre_m][n]['dist'] + G.edge[n][suc_m]['dist']
                            dist_m_prop = G.edge[pre_n][m]['dist'] + G.edge[m][suc_n]['dist']
                            if dist_n_prop + dist_m_prop < dist_n_ori +dist_m_ori:
                                print str()+'meta-heuristic swapping executed between: ',(n,m)
                                G_c.add_edge(pre_m, n, attr_dict = G.get_edge_data(pre_m, n))
                                G_c.add_edge(n, suc_m, attr_dict = G.get_edge_data(n, suc_m))
                                G_c.add_edge(pre_n, m, attr_dict = G.get_edge_data(pre_n, m))
                                G_c.add_edge(m, suc_n, attr_dict = G.get_edge_data(m, suc_n))
                                
                                G_c.add_edge(pre_m, n, attr_dict = G.get_edge_data(pre_m, n))
                                G_c.add_edge(n, suc_m, attr_dict = G.get_edge_data(n, suc_m))
                                G_c.add_edge(pre_n, m, attr_dict = G.get_edge_data(pre_n, m))
                                G_c.add_edge(m, suc_n, attr_dict = G.get_edge_data(m, suc_n))
                                

                                if suc_n == suc_m:
                                    G_c.remove_edge(pre_m,m)
                                    G_c.remove_edge(pre_n,n)
                                else:
                                    if pre_n == pre_m:
                                        G_c.remove_edge(m,suc_m)
                                        G_c.remove_edge(n,suc_n)
                                    else:
                                        G_c.remove_edge(pre_m,m)
                                        G_c.remove_edge(pre_n,n)
                                        G_c.remove_edge(n,suc_n)
                                        G_c.remove_edge(m,suc_m)

                        

    print avg_speed[2:-1]
    solution = transfer_graph_to_solution_dic(G_c)
    visited = set()
    t_c_number = {}
    
    for t,p in solution.iteritems():
        t_c_number[t] = len(p)
        for n in p:
            for i in n:
                visited.add(i)

    t_nmbr = 0
    for t in t_c_number:
        t_nmbr += t_c_number[t]
    
    return_list = [solution, 
                   {'terminals':len(solution.keys()),
                    'Title':'W.E. Swapping', 
                    'conductor number':t_nmbr, 
                    'remained W.E.':len(G.nodes())-len(visited)
                    }
                   ]
    return return_list


def avg_speed(G, sol_dic):
    G_c = G.copy()
    routes_in_edge = []
    solution_l = []
    for k,v in sol_dic.iteritems():
        for r in v:
            solution_l.append(r)
    
    for p in solution_l:
        routes_in_edge.append([(p[n],p[n+1]) for n in range(len(p)-1)])   

    ttl_dist = 0
    ttl_time = 0
    for p in routes_in_edge:
        for e in p:
            ttl_dist += G_c.edge[e[0]][e[1]]['dist']
            ttl_time += G_c.edge[e[0]][e[1]]['delt_t']
    
    avg_spd = ttl_dist/ttl_time
    return avg_spd

def plot_t_x_plot(G_origin, solu, dot_line=False, plt_ttl = True, width = 1200, height = 900):
    attr_dic = solu[1]
    routes_in_edge = []
    solution_l = []
    for k,v in solu[0].iteritems():
        for r in v:
            solution_l.append(r)
    for p in solution_l:
        routes_in_edge.append([(p[n],p[n+1]) for n in range(len(p)-1)])
    
    colors =        ['r',   'b',    'y',    'g',    'c',    'm',    'k']
    linewidths =    [5 for n in routes_in_edge]
    
    
    node_label_dict = {}
    pos_dict = {}
    
    for (n,attr) in G_origin.nodes(data=True):
        (node_label_dict[n], pos_dict[n]) = (attr['label'], attr['coord'])
    
    
    G_draw_edge_label = G_origin.copy()
    if dot_line:
    #-------plot the dashlines----#
        nx.draw_networkx(G_origin,
                         labels =node_label_dict,
                         edgelist = G.edges(), 
                         pos = pos_dict, 
                         edge_color = 'purple', 
                         style = 'dashed', 
                         node_color = 'pink',
                         arrow = False
                         )
    #-------plot the dashlines----#
    
    #-------plot the conductor trips----#
    for ct, route in enumerate(routes_in_edge):
       
        if route[0][0] == 'T0Begin':
            color = 'y'
        if route[0][0] == 'T1Begin':
            color = 'b'
        if route[0][0] == 'T2Begin':
            color = 'r'
        if route[0][0] == 'T3Begin':
            color = 'k'
        
        nx.draw_networkx(G, 
                         labels = node_label_dict, 
                         edgelist = route, 
                         pos = pos_dict, 
                         width = linewidths[ct],
                         edge_color = color,
                         alpha = 0.5, 
                         arrows = False)
    #-------plot the conductor trips----#
    #-------plot edge speed-------------#
    edges_to_draw = []
    
    for r in routes_in_edge:
        for e in r:
            edges_to_draw.append(e)
    
    for e in G_draw_edge_label.edges():
        if e not in edges_to_draw:
            G_draw_edge_label.remove_edge(e[0],e[1])
    plt.get_current_fig_manager().resize(width, height)
    spd_label = nx.get_edge_attributes(G_draw_edge_label, 'spd')
    
    nx.draw_networkx_edge_labels(G_draw_edge_label,  
                                 pos=pos_dict,
                                 edge_labels = spd_label, 
                                 alpha = 15.0)
    
    
    
    
    if plt_ttl:
        ttl = ''
        for k,v in attr_dic.iteritems():
            if k == 'Title':
                ttl = ttl + k +':  ' + str(v) +';'
            if k == 'terminals':
                plt.text(terminal_loc[-1]/2, shift_hour+0.5, k+':  '+str(v))
            if k == 'conductor number':
                plt.text(terminal_loc[-1]/2, shift_hour+1, k+':  '+str(v))
            if k == 'remained W.E.':
                plt.text(terminal_loc[-1]/2, shift_hour+1.5, k+':  '+str(v))
            
        plt.title(ttl)
    
            
    
    plt.show()
    
    #-------plot edge speed-------------#
    
    
    #------------play with parameters here------------#
max_speed = 15
terminal_loc = [0,100,200,300]
shift_hour = 12
data = data_init(sample=True, WE_number=35)
G = graph_create(data, terminal_loc = terminal_loc, shift_hr = shift_hour)
def monte_carlo(n = 10):
    spd_list = []
    solution_list = []
    for i in range(n):
        solution = construct_heu_feasi_sol(G, max_speed=max_speed, dist_first=False, rand = True)
        solution = path_release(G, sol_dic=solution[0], max_speed=max_speed, release_condition=1)   
        solution = meta_heu(G, sol_dic=solution[0], max_speed=max_speed)
        avg_spd = avg_speed(G, sol_dic=solution[0])
        #------------play with parameters here------------#
        
        spd_list.append(avg_spd)
        solution_list.append(solution)
    best_sol = solution_list[spd_list.index(min(spd_list))]
    return best_sol

solution = monte_carlo()


visited = set()
t_c_number = {}
for t,p in solution[0].iteritems():
    t_c_number[t] = len(p)
    for n in p:
        for i in n:
            visited.add(i)
        
t_nmbr = 0

for t in t_c_number:
    print '  number of feasible solutions for terminal_'+str(t)+' is: '+str(t_c_number[t])
    t_nmbr += t_c_number[t]
print '  number of terminals: '+ str(len(t_c_number.keys()))
print '  number of total solutions: '+ str(t_nmbr)
print '  number of remained W.E.: '+str(len(G.nodes())-len(visited))


endt = time.time()
print ('runtime: '+ str(endt-startt))
print ('program ends')


plot_t_x_plot(G, solu=solution, dot_line=False, plt_ttl = True, width = 1600, height=1200)

